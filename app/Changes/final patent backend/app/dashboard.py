"""Plan Dashboard-sheet cell updates from computed stats.

Cell positions differ slightly between the workbooks, so labels are located
dynamically in a downloaded copy and the matching cells are patched in-place
via the Graph workbook API (formatting, charts and everything else untouched).
"""
import io
from datetime import date

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .dedupe import (
    ALLOWANCE,
    GRANTED,
    NAT_PHASE,
    NON_PROV,
    PCT_FILED,
    PCT_VIA_NP,
    PROV,
    Stats,
    norm_text,
)

# normalized label -> stats key ("total"/"rejected" are specials)
_SUMMARY_LABELS = {
    "total patents": "total",
    "granted": GRANTED,
    "pct filed": PCT_FILED,
    "pct via non provisional": PCT_VIA_NP,
    "non provisional": NON_PROV,
    "provisional": PROV,
    "national phase": NAT_PHASE,
    "rejected patents": "rejected",
    "allowance issued": ALLOWANCE,
}

_STATUS_LABELS = {
    "granted": GRANTED,
    "pct application filed": PCT_FILED,
    "pct via non provisional": PCT_VIA_NP,
    "non provisional submitted": NON_PROV,
    "provisional submitted": PROV,
    "national phase application": NAT_PHASE,
    "allowance issued": ALLOWANCE,
}

ALLOWANCE_LIST_ROWS = 5
DEADLINE_LIST_ROWS = 5


def _addr(row: int, col: int, row2: int | None = None, col2: int | None = None) -> str:
    a = f"{get_column_letter(col)}{row}"
    if row2 is None:
        return a
    return f"{a}:{get_column_letter(col2 or col)}{row2}"


def _stat_value(stats: Stats, key: str):
    if key == "total":
        return stats.total
    if key == "rejected":
        return stats.rejected
    return stats.status_counts.get(key, 0)


def plan_dashboard_updates(
    xlsx_bytes: bytes,
    stats: Stats,
    sheet: str = "Dashboard",
    per_source: list[tuple[str, Stats]] | None = None,
    unlicensed_granted: int = 0,
    per_source_unlicensed: dict[str, int] | None = None,
) -> list[tuple[str, list[list]]]:
    """Return [(range_address, values_2d), ...] to patch on the Dashboard sheet.

    per_source: [(source_label, stats), ...] — fills the master's
    Per-Source Summary / Per-Source Category Breakdown tables when present.
    unlicensed_granted: N granted-but-unlicensed patents shown as "+NP" in the
    top summary's Total and Granted cells (e.g. granted 8 -> "5+3P").
    per_source_unlicensed: {source_label: N} — per-source unlicensed granted
    counts used to format individual source rows in the Per-Source Summary
    (e.g. Purdue, K&K granted 6 -> "3+3P").
    """
    per_source_unlicensed = per_source_unlicensed or {}
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    ws = wb[sheet]
    cells: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                cells[(c.row, c.column)] = c.value

    def find(pred):
        for (r, c), v in cells.items():
            if pred(v):
                return r, c
        return None

    updates: list[tuple[str, list[list]]] = []
    total = stats.total or 1  # avoid div-by-zero in percentages

    # --- top summary row: labels somewhere in a row, values in the row below ---
    pos = find(lambda v: norm_text(v) == "total patents")
    if pos:
        r, c0 = pos
        row_labels = [(c, norm_text(v)) for (rr, c), v in cells.items() if rr == r]
        mapped = {c: _SUMMARY_LABELS[t] for c, t in row_labels if t in _SUMMARY_LABELS}
        cmin, cmax = min(mapped), max(mapped)
        row_values = []
        for c in range(cmin, cmax + 1):
            if c not in mapped:
                row_values.append(None)
                continue
            v = _stat_value(stats, mapped[c])
            if unlicensed_granted and mapped[c] in ("total", GRANTED):
                v = f"{max(v - unlicensed_granted, 0)}+{unlicensed_granted}P"
            row_values.append(v)
        updates.append((_addr(r + 1, cmin, r + 1, cmax), [row_values]))

    # --- Status Breakdown table ---
    pos = find(lambda v: norm_text(v) == "status breakdown")
    if pos:
        r0, c0 = pos
        for r in range(r0 + 2, r0 + 12):
            label = norm_text(cells.get((r, c0)))
            if not label:
                break
            key = _STATUS_LABELS.get(label)
            if key is None:
                continue
            count = stats.status_counts.get(key, 0)
            shown = count
            if unlicensed_granted and key == GRANTED:
                shown = f"{max(count - unlicensed_granted, 0)}+{unlicensed_granted}P"
            updates.append((_addr(r, c0 + 1, r, c0 + 2), [[shown, count / total]]))

    # --- Category Breakdown table ---
    pos = find(lambda v: norm_text(v) == "category breakdown")
    if pos:
        r0, c0 = pos
        for r in range(r0 + 2, r0 + 10):
            label = cells.get((r, c0))
            n = norm_text(label)
            if not n:
                break
            match = next(
                (cat for cat, cnt in stats.category_counts.items() if norm_text(cat).split(" ")[0] == n.split(" ")[0]),
                None,
            )
            count = stats.category_counts.get(match, 0) if match else 0
            updates.append((_addr(r, c0 + 1, r, c0 + 2), [[count, count / total]]))

    # --- Action Items alert count ---
    pos = find(lambda v: norm_text(v).startswith("patents ready for grant"))
    if pos:
        r, c = pos
        updates.append((_addr(r + 1, c), [[len(stats.allowance_ready)]]))

    # --- Allowance list (Title / Deadline / Ref Number) ---
    pos = find(lambda v: norm_text(v) == "patents with notice of allowance issued")
    if pos:
        r, c = pos
        rows = []
        for i in range(ALLOWANCE_LIST_ROWS):
            if i < len(stats.allowance_ready):
                a = stats.allowance_ready[i]
                rows.append([a["title"], a["deadline"], a["ref"]])
            else:
                rows.append(["", "", ""])
        updates.append((_addr(r + 2, c, r + 1 + ALLOWANCE_LIST_ROWS, c + 2), rows))

    # --- Upcoming deadlines (Title / Deadline / Status) ---
    pos = find(lambda v: "upcoming deadlines" in norm_text(v))
    if pos:
        r, c = pos
        rows = []
        for i in range(DEADLINE_LIST_ROWS):
            if i < len(stats.upcoming_deadlines):
                d = stats.upcoming_deadlines[i]
                rows.append([d["title"], d["deadline"], d["status"]])
            else:
                rows.append(["", "", ""])
        updates.append((_addr(r + 2, c, r + 1 + DEADLINE_LIST_ROWS, c + 2), rows))

    # --- Per-source tables (master only) ---
    if per_source:
        pos = find(lambda v: norm_text(v) == "per source summary")
        if pos:
            r0, c0 = pos
            header_row = r0 + 1
            col_labels = {c: norm_text(v) for (rr, c), v in cells.items() if rr == header_row}
            old_rows = 0
            while cells.get((header_row + 1 + old_rows, c0)) not in (None, ""):
                old_rows += 1
            rows = []
            cmin, cmax = min(col_labels), max(col_labels)
            for label, s in per_source:
                src_unlicensed = per_source_unlicensed.get(label, 0)
                row = []
                for c in range(cmin, cmax + 1):
                    t = col_labels.get(c, "")
                    if t == "source":
                        row.append(label)
                    elif t == "total":
                        v = s.total
                        if src_unlicensed:
                            v = f"{max(v - src_unlicensed, 0)}+{src_unlicensed}P"
                        row.append(v)
                    elif t == "rejected patents":
                        row.append(s.rejected)
                    elif t.startswith("pct via"):
                        row.append(s.status_counts.get(PCT_VIA_NP, 0))
                    elif t in _SUMMARY_LABELS:
                        v = _stat_value(s, _SUMMARY_LABELS[t])
                        if src_unlicensed and _SUMMARY_LABELS[t] == GRANTED:
                            v = f"{max(v - src_unlicensed, 0)}+{src_unlicensed}P"
                        row.append(v)
                    else:
                        row.append(None)
                rows.append(row)
            while len(rows) < old_rows:
                rows.append(["" for _ in range(cmin, cmax + 1)])
            updates.append((_addr(header_row + 1, cmin, header_row + len(rows), cmax), rows))

        pos = find(lambda v: norm_text(v) == "per source category breakdown")
        if pos:
            r0, c0 = pos
            header_row = r0 + 1
            col_labels = {c: norm_text(v) for (rr, c), v in cells.items() if rr == header_row}
            old_rows = 0
            while cells.get((header_row + 1 + old_rows, c0)) not in (None, ""):
                old_rows += 1
            cmin, cmax = min(col_labels), max(col_labels)
            rows = []
            for label, s in per_source:
                row = []
                for c in range(cmin, cmax + 1):
                    t = col_labels.get(c, "")
                    if t == "source":
                        row.append(label)
                    elif t:
                        match = next(
                            (cat for cat in s.category_counts if norm_text(cat).split(" ")[0] == t.split(" ")[0]),
                            None,
                        )
                        row.append(s.category_counts.get(match, 0) if match else 0)
                    else:
                        row.append(None)
                rows.append(row)
            while len(rows) < old_rows:
                rows.append(["" for _ in range(cmin, cmax + 1)])
            updates.append((_addr(header_row + 1, cmin, header_row + len(rows), cmax), rows))

    # --- "Last updated" line: only when static text (the sources use a live formula) ---
    pos = find(lambda v: isinstance(v, str) and v.startswith("Live Summary"))
    if pos:
        r, c = pos
        updates.append(
            (_addr(r, c), [[f"Live Summary - Last updated: {date.today().strftime('%d-%b-%Y')}"]])
        )

    return updates
