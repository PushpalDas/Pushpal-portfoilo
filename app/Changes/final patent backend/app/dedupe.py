"""Parse Patents sheets, group duplicate/continuation rows, compute dashboard stats.

Dedup rule (from the team): a patent may appear as several consecutive rows —
either the same title repeated or a blank-title row directly under its parent
(e.g. a PCT filing following the non-provisional). Each group counts as ONE
patent and the status of the LAST row in the group is the patent's status.
"""
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# canonical status keys
GRANTED = "Granted"
PCT_FILED = "PCT application filed"
PCT_VIA_NP = "PCT Via Non provisional"
NON_PROV = "Non provisional submitted"
PROV = "Provisional submitted"
NAT_PHASE = "National phase application"
ALLOWANCE = "Allowance issued"

_STATUS_CANON = {
    "granted": GRANTED,
    "pct application filed": PCT_FILED,
    "pct filed": PCT_FILED,
    "pct via non provisional": PCT_VIA_NP,
    "non provisional submitted": NON_PROV,
    "non provsional submitted": NON_PROV,
    "provisional submitted": PROV,
    "national phase application": NAT_PHASE,
    "allowance issued": ALLOWANCE,
}

CATEGORIES = ["Comms", "Compute", "Sensing", "Security", "Applications (combo of many of these)"]


def norm_text(v) -> str:
    """Normalize label/status text: unicode dashes/spaces -> ascii, lowercase, collapse."""
    if v is None:
        return ""
    s = str(v)
    for ch in ("‑", "–", "—"):
        s = s.replace(ch, "-")
    for ch in (" ", "\xa0"):
        s = s.replace(ch, " ")
    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def canon_status(v) -> str | None:
    n = norm_text(v)
    if not n:
        return None
    return _STATUS_CANON.get(n, str(v).strip())


def _clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.replace("_x000D_", " ").strip()
        return s or None
    return v


def _as_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_patents(xlsx_bytes: bytes, sheet: str = "Patents") -> tuple[list[str], list[dict]]:
    """Return (headers, rows) of the data table on the Patents sheet."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[sheet]
    if ws.tables:
        ref = next(iter(ws.tables.values())).ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
    headers = []
    for c in range(min_col, max_col + 1):
        h = ws.cell(min_row, c).value
        headers.append(str(h).strip() if h is not None else f"col{c}")
    rows = []
    for r in range(min_row + 1, max_row + 1):
        vals = {h: _clean(ws.cell(r, c).value) for h, c in zip(headers, range(min_col, max_col + 1))}
        if any(v is not None for v in vals.values()):
            rows.append(vals)
    return headers, rows


@dataclass
class Patent:
    title: str
    status: str | None          # canonical status of the LAST row of the group
    category: str | None
    rejected: bool
    action_items: str | None    # last non-blank Action Items
    actions: list[str]          # Action Items from ALL rows of the group
    ref: str | None             # last non-blank Ref Number
    deadlines: list[date] = field(default_factory=list)
    row_count: int = 1


def dedupe(rows: list[dict]) -> list[Patent]:
    groups: dict[str, list[dict]] = {}
    order: list[str] = []
    last_key = None
    for i, row in enumerate(rows):
        title = row.get("Title")
        if title:
            key = norm_text(title)
        elif last_key is not None:
            key = last_key  # blank-title continuation row belongs to the patent above
        else:
            key = f"__row{i}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)
        last_key = key

    patents = []
    for key in order:
        grp = groups[key]
        title = next((r["Title"] for r in grp if r.get("Title")), "")
        status = None
        for r in reversed(grp):  # status of the latest (last) row wins
            if r.get("Status"):
                status = canon_status(r["Status"])
                break
        category = next((r["Category"] for r in reversed(grp) if r.get("Category")), None)
        rejected = any(norm_text(r.get("Rejected Patent")) == "yes" for r in grp)
        action = next((r["Action Items"] for r in reversed(grp) if r.get("Action Items")), None)
        actions = [str(r["Action Items"]) for r in grp if r.get("Action Items")]
        ref = next((str(r["Ref Number"]) for r in reversed(grp) if r.get("Ref Number")), None)
        deadlines = [d for d in (_as_date(r.get("Deadlines")) for r in grp) if d]
        patents.append(
            Patent(
                title=str(title),
                status=status,
                category=str(category) if category is not None else None,
                rejected=rejected,
                action_items=str(action) if action is not None else None,
                actions=actions,
                ref=ref,
                deadlines=deadlines,
                row_count=len(grp),
            )
        )
    return patents


@dataclass
class Stats:
    total: int
    status_counts: dict[str, int]
    rejected: int
    category_counts: dict[str, int]
    allowance_ready: list[dict]      # patents whose action items mention notice of allowance
    upcoming_deadlines: list[dict]   # next 5 deadlines from today

    def as_dict(self) -> dict:
        return {
            "total": self.total,
            "status_counts": self.status_counts,
            "rejected": self.rejected,
            "category_counts": self.category_counts,
            "allowance_ready_count": len(self.allowance_ready),
            "allowance_ready": self.allowance_ready,
            "upcoming_deadlines": self.upcoming_deadlines,
        }


def compute_stats(patents: list[Patent], today: date | None = None) -> Stats:
    today = today or date.today()
    status_counts: dict[str, int] = {
        s: 0 for s in (GRANTED, PCT_FILED, PCT_VIA_NP, NON_PROV, PROV, NAT_PHASE, ALLOWANCE)
    }
    for p in patents:
        if p.status:
            status_counts[p.status] = status_counts.get(p.status, 0) + 1

    category_counts = {c: 0 for c in CATEGORIES}
    for p in patents:
        if not p.category:
            continue
        n = norm_text(p.category)
        for c in CATEGORIES:
            if norm_text(c).split(" ")[0] == n.split(" ")[0]:
                category_counts[c] += 1
                break

    # a patent is "ready for grant" if ANY of its rows mentions a notice of
    # allowance in Action Items (not just the latest row)
    allowance_ready = [
        {
            "title": p.title,
            "deadline": max(p.deadlines).isoformat() if p.deadlines else "",
            "ref": p.ref or "",
        }
        for p in patents
        if any("notice of allowance issued" in norm_text(a) for a in p.actions)
    ]

    upcoming = sorted(
        (
            {"title": p.title, "deadline": d.isoformat(), "status": p.status or ""}
            for p in patents
            for d in p.deadlines
            if d >= today
        ),
        key=lambda x: x["deadline"],
    )[:5]

    return Stats(
        total=len(patents),
        status_counts=status_counts,
        rejected=sum(1 for p in patents if p.rejected),
        category_counts=category_counts,
        allowance_ready=allowance_ready,
        upcoming_deadlines=upcoming,
    )


def combine_stats(stats_list: list[Stats], today: date | None = None) -> Stats:
    """Aggregate per-source stats into the master's combined stats."""
    status_counts: dict[str, int] = {}
    category_counts: dict[str, int] = {}
    allowance = []
    deadlines = []
    for s in stats_list:
        for k, v in s.status_counts.items():
            status_counts[k] = status_counts.get(k, 0) + v
        for k, v in s.category_counts.items():
            category_counts[k] = category_counts.get(k, 0) + v
        allowance.extend(s.allowance_ready)
        deadlines.extend(s.upcoming_deadlines)
    return Stats(
        total=sum(s.total for s in stats_list),
        status_counts=status_counts,
        rejected=sum(s.rejected for s in stats_list),
        category_counts=category_counts,
        allowance_ready=allowance,
        upcoming_deadlines=sorted(deadlines, key=lambda x: x["deadline"])[:5],
    )
