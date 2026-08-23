"""Rebuild the master workbook's Patents sheet from the three source workbooks."""
import io
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

MASTER_SHEET = "Patents"


def _cell_value(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        return v.replace("_x000D_", " ")
    return v


def plan_master_patents_update(
    master_bytes: bytes,
    source_rows: list[tuple[str, list[dict]]],
) -> list[tuple[str, list[list]]]:
    """Build range updates that rewrite the master Patents table in its current
    structure: same header columns (Source first), one row per source row.

    source_rows: [(source_label, rows_from_that_source), ...] in write order.
    """
    wb = load_workbook(io.BytesIO(master_bytes), data_only=False)
    ws = wb[MASTER_SHEET]
    if ws.tables:
        ref = next(iter(ws.tables.values())).ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    else:
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row

    headers = []
    for c in range(min_col, max_col + 1):
        h = ws.cell(min_row, c).value
        headers.append(str(h).strip() if h is not None else "")

    data: list[list] = []
    for label, rows in source_rows:
        for row in rows:
            out = []
            for h in headers:
                if h.lower() == "source":
                    out.append(label)
                else:
                    out.append(_cell_value(row.get(h)))
            data.append(out)

    # pad with blank rows to clear anything left over from the previous state;
    # scan actual cell values (ws.max_row includes formatting-only rows)
    last_data_row = min_row
    for r in range(min_row + 1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(min_col, max_col + 1)):
            last_data_row = r
    old_rows = max(last_data_row, max_row) - min_row
    while len(data) < old_rows:
        data.append([""] * len(headers))

    start = min_row + 1
    end = min_row + len(data)
    address = f"{get_column_letter(min_col)}{start}:{get_column_letter(max_col)}{end}"
    return [(address, data)]
